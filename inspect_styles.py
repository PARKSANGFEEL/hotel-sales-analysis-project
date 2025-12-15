import openpyxl
from openpyxl.styles import PatternFill

file_path = '매출상세내역 목록_20251207_final_v16.xlsx'
sheet_name = '점유율_및_평균단가_v5'

try:
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        print(f"Sheet '{sheet_name}' not found.")
        exit()
        
    ws = wb[sheet_name]
    
    print(f"Inspecting sheet: {sheet_name}")
    
    # Find occupancy columns (Row 2)
    occupancy_cols = {}
    header_row = ws[2]
    for cell in header_row:
        if cell.value and '점유율(%)' in str(cell.value):
            occupancy_cols[cell.column] = cell.value
            
    print(f"Occupancy columns found: {occupancy_cols}")
    
    # Check a few rows (Start from Row 3)
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=20), start=3):
        for col_idx, cell in enumerate(row, start=1):
            if col_idx in occupancy_cols:
                value = cell.value
                fill = cell.fill
                
                color_info = "No Fill"
                if fill and fill.fill_type == 'solid':
                    if hasattr(fill.start_color, 'rgb'):
                        color_code = fill.start_color.rgb
                        # openpyxl often prepends 00 for alpha if not specified, or just uses the hex
                        if color_code in ['FFFFFF00', '00FFFF00', 'FFFF00']:
                            color_info = "YELLOW"
                        elif color_code in ['FFFF0000', '00FF0000', 'FF0000']:
                            color_info = "RED"
                        elif color_code in ['FFFFA500', '00FFA500', 'FFA500']:
                            color_info = "ORANGE"
                        else:
                            color_info = f"Color: {color_code}"
                
                if isinstance(value, (int, float)):
                    print(f"Row {row_idx}, Col {col_idx} ({occupancy_cols[col_idx]}): Value {value} -> {color_info}")
        for col_idx, cell in enumerate(row, start=1):
            if col_idx in occupancy_cols:
                value = cell.value
                fill = cell.fill
                
                color_info = "No Fill"
                if fill and fill.fill_type == 'solid':
                    # start_color is usually an object with 'rgb' or 'index'
                    # For PatternFill created with hex, it's usually in rgb
                    if hasattr(fill.start_color, 'rgb'):
                        color_code = fill.start_color.rgb
                        if color_code == 'FFFFFF00':
                            color_info = "YELLOW"
                        elif color_code == 'FFFF0000':
                            color_info = "RED"
                        else:
                            color_info = f"Color: {color_code}"
                
                if isinstance(value, (int, float)):
                    if value >= 80 and color_info != "YELLOW":
                        print(f"Row {row_idx}, Col {col_idx} ({occupancy_cols[col_idx]}): Value {value} -> Expected YELLOW, got {color_info}")
                    elif value <= 50 and color_info != "RED":
                        print(f"Row {row_idx}, Col {col_idx} ({occupancy_cols[col_idx]}): Value {value} -> Expected RED, got {color_info}")
                    elif (value > 50 and value < 80) and color_info != "No Fill":
                         print(f"Row {row_idx}, Col {col_idx} ({occupancy_cols[col_idx]}): Value {value} -> Expected No Fill, got {color_info}")
                    else:
                        # Correct case, maybe print a few examples
                        if row_idx <= 5:
                             print(f"Row {row_idx}, Col {col_idx} ({occupancy_cols[col_idx]}): Value {value} -> Correctly {color_info}")

except Exception as e:
    print(f"Error: {e}")
