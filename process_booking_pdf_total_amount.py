import pdfplumber
import pandas as pd
import glob
import os
import re
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from dateutil.relativedelta import relativedelta

# Configuration
pdf_folder = r'c:\Users\HP\Desktop\호텔\부킹매출자료'
excel_file = '매출상세내역 목록_20251207_complete.xlsx'

def parse_pdf_line(line):
    parts = line.split()
    if len(parts) < 12 or parts[0] != '예약' or 'KRW' not in parts:
        return None
    
    try:
        krw_index = -1
        for i in range(len(parts) - 1, -1, -1):
            if parts[i] == 'KRW':
                krw_index = i
                break
        
        if krw_index == -1:
            return None

        amount_str = parts[krw_index + 1]
        amount = float(amount_str.replace(',', ''))
        
        check_in_str = f"{parts[3]} {parts[4]} {parts[5]}"
        check_out_str = f"{parts[6]} {parts[7]} {parts[8]}"
        
        name_parts = parts[9:krw_index]
        name = " ".join(name_parts)
        
        return {
            '체크인': check_in_str,
            '체크아웃': check_out_str,
            '고객명': name,
            '금액': amount
        }
    except Exception:
        return None

def extract_payment_date(text):
    match = re.search(r'대금 지급 날짜\s+(\d{4}년\s+\d{1,2}월\s+\d{1,2}일)', text)
    if match:
        return match.group(1)
    return None

def extract_total_amount(text):
    # Pattern: 대금 총액 ₩69,519,200
    match = re.search(r'대금 총액\s+₩([\d,]+)', text)
    if match:
        return float(match.group(1).replace(',', ''))
    return None

def process_booking_data():
    all_details = []
    monthly_summary_data = []
    
    pdf_files = glob.glob(os.path.join(pdf_folder, '*.pdf'))
    print(f"Found {len(pdf_files)} PDF files.")
    
    for pdf_file in pdf_files:
        pdf_payment_date = None
        pdf_total_amount = 0
        pdf_details = []
        
        try:
            with pdfplumber.open(pdf_file) as pdf:
                # 1. Extract Payment Date (usually on Page 1)
                first_page_text = pdf.pages[0].extract_text()
                pdf_payment_date = extract_payment_date(first_page_text)
                
                # 2. Extract Total Amount (usually on Last Page)
                last_page_text = pdf.pages[-1].extract_text()
                pdf_total_amount = extract_total_amount(last_page_text)
                
                if pdf_total_amount is None:
                    print(f"Warning: Total Amount not found in {os.path.basename(pdf_file)}. Calculating from rows.")
                    pdf_total_amount = 0 # Will be updated by row sum if not found
                    use_row_sum = True
                else:
                    use_row_sum = False

                # 3. Extract Details
                for page in pdf.pages:
                    text = page.extract_text()
                    if not text: continue
                    
                    lines = text.split('\n')
                    for line in lines:
                        if line.startswith('예약'):
                            parsed = parse_pdf_line(line)
                            if parsed:
                                pdf_details.append(parsed)
                                if use_row_sum:
                                    pdf_total_amount += parsed['금액']
            
            all_details.extend(pdf_details)
            
            # Determine Sales Month (Payment Date - 1 Month)
            if pdf_payment_date:
                date_str = pdf_payment_date.replace('년', '-').replace('월', '-').replace('일', '').replace(' ', '')
                try:
                    dt = pd.to_datetime(date_str, format='%Y-%m-%d')
                    # Subtract 1 month
                    sales_month_dt = dt - relativedelta(months=1)
                    month_str = sales_month_dt.strftime('%Y-%m')
                except:
                    month_str = pdf_payment_date
            else:
                month_str = "Unknown"
                
            monthly_summary_data.append({
                '월': month_str,
                '금액': pdf_total_amount,
                '파일명': os.path.basename(pdf_file)
            })
            print(f"Processed {os.path.basename(pdf_file)}: Month={month_str}, Total={pdf_total_amount}")
            
        except Exception as e:
            print(f"Error processing {pdf_file}: {e}")

    if not all_details:
        print("No booking data found.")
        return

    # 1. Booking Sales Sheet (Details)
    df_details = pd.DataFrame(all_details)
    def clean_date(d):
        return d.replace('년', '-').replace('월', '-').replace('일', '').replace(' ', '')
    df_details['체크인'] = df_details['체크인'].apply(clean_date)
    df_details['체크아웃'] = df_details['체크아웃'].apply(clean_date)
    
    booking_sales_df = df_details[['체크인', '체크아웃', '고객명', '금액']].copy()
    
    # Add Total Row for Details
    total_amount_details = booking_sales_df['금액'].sum()
    total_row = pd.DataFrame([{'체크인': '합계', '체크아웃': '', '고객명': '', '금액': total_amount_details}])
    booking_sales_df_with_total = pd.concat([booking_sales_df, total_row], ignore_index=True)
    
    # 2. Monthly Booking Sales Data (From PDF Totals)
    df_monthly = pd.DataFrame(monthly_summary_data)
    # Group by Month
    df_monthly_grouped = df_monthly.groupby('월')['금액'].sum().reset_index()
    
    # Add Total Row for Monthly
    monthly_total = df_monthly_grouped['금액'].sum()
    monthly_total_row = pd.DataFrame([{'월': '합계', '금액': monthly_total}])
    monthly_booking_sales_with_total = pd.concat([df_monthly_grouped, monthly_total_row], ignore_index=True)

    # Save to Excel
    try:
        wb = openpyxl.load_workbook(excel_file)
        
        def write_sheet(wb, name, data):
            if name in wb.sheetnames:
                print(f"Sheet '{name}' already exists. Overwriting...")
                del wb[name]
            ws = wb.create_sheet(name)
            for r in dataframe_to_rows(data, index=False, header=True):
                ws.append(r)
            return ws

        write_sheet(wb, '부킹매출', booking_sales_df_with_total)
        write_sheet(wb, '부킹_월별매출', monthly_booking_sales_with_total)
        
        # Apply styles
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        comma_format = '#,##0'
        
        for sheet_name in ['부킹매출', '부킹_월별매출']:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row > 1:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = comma_format
        
        wb.save(excel_file)
        print(f"Successfully updated booking data in {excel_file} (Using PDF Total Amount)")
        
    except Exception as e:
        print(f"Error saving to Excel: {e}")

if __name__ == "__main__":
    process_booking_data()
