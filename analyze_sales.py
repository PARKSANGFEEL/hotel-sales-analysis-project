import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font
import glob
import os

# 매출상세내역 목록으로 시작하는 파일 자동 검색 (analyzed 파일 제외)
file_pattern = '매출상세내역 목록*.xlsx'
matching_files = glob.glob(file_pattern)

# analyzed 파일 제외
matching_files = [f for f in matching_files if '_analyzed' not in f]

if not matching_files:
    raise FileNotFoundError(f"'{file_pattern}' 패턴과 일치하는 파일을 찾을 수 없습니다.")

# 가장 최근에 수정된 파일 선택
file_path = max(matching_files, key=os.path.getmtime)
print(f"사용할 파일: {file_path}")

# 출력 파일명 생성 (원본 파일명 기반)
base_name = os.path.splitext(file_path)[0]
output_file_path = f"{base_name}_analyzed.xlsx"

# 시트 이름 자동 확인
wb_temp = openpyxl.load_workbook(file_path)
available_sheets = wb_temp.sheetnames
wb_temp.close()

# 가능한 시트 이름 목록 (우선순위 순)
possible_sheet_names = ['매출정리', '상세 빌 목록', 'Sheet1']
source_sheet = None

for sheet_name in possible_sheet_names:
    if sheet_name in available_sheets:
        source_sheet = sheet_name
        break

if source_sheet is None:
    source_sheet = available_sheets[0]  # 첫 번째 시트 사용

print(f"사용할 시트: {source_sheet}")

try:
    # Load the source data
    df = pd.read_excel(file_path, sheet_name=source_sheet)
    
    # 컬럼명 매핑 (다양한 파일 형식 지원)
    column_mapping = {
        '객실/단체ID': '객실',
        '객실/단체 ID': '객실'
    }
    
    # 컬럼명 변경
    df.rename(columns=column_mapping, inplace=True)
    
    # '객실' 컬럼이 없으면 에러
    if '객실' not in df.columns:
        available_cols = ', '.join(df.columns.tolist())
        raise ValueError(f"'객실' 또는 '객실/단체ID' 컬럼을 찾을 수 없습니다. 사용 가능한 컬럼: {available_cols}")
    
    # Remove the '합계' and '소계' rows if they exist
    # Filter out summary rows that don't represent actual room numbers
    df = df[df['객실'] != '합계']
    df = df[df['객실'] != '소계']
    
    # Remove rows where 객실 is NaN or empty
    df = df[df['객실'].notna()]
    df = df[df['객실'].astype(str).str.strip() != '']
    
    # Keep only rows where 객실 looks like a room number (starts with digits)
    df = df[df['객실'].astype(str).str.match(r'^\d', na=False)]
    
    # Ensure numeric columns are numbers
    numeric_cols = ['객실료', '부가세', '총금액']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    # Convert '일자' to datetime and extract Month
    df['일자'] = pd.to_datetime(df['일자'], errors='coerce')
    # Remove rows where date conversion failed or is NaT
    df = df.dropna(subset=['일자'])
    df['월'] = df['일자'].dt.strftime('%Y-%m')

    # 1. Monthly Sales
    monthly_sales = df.groupby('월')[numeric_cols].sum().reset_index()
    # Add total row for monthly sales
    total_row = pd.DataFrame([['합계'] + [monthly_sales[col].sum() for col in numeric_cols]], 
                             columns=['월'] + numeric_cols)
    monthly_sales = pd.concat([monthly_sales, total_row], ignore_index=True)
    
    # 2. Room Sales
    room_sales = df.groupby('객실')[numeric_cols].sum().reset_index()
    # Add total row for room sales
    total_row = pd.DataFrame([['합계'] + [room_sales[col].sum() for col in numeric_cols]], 
                             columns=['객실'] + numeric_cols)
    room_sales = pd.concat([room_sales, total_row], ignore_index=True)
    
    # 3. Monthly & Room Sales (Pivot)
    # Using pivot_table to show Rooms as rows and Months as columns for Total Amount
    monthly_room_sales = df.pivot_table(index='객실', columns='월', values='총금액', aggfunc='sum', fill_value=0).reset_index()
    
    # Add row total (sum across all months for each room) - 마지막 열에 합계
    month_cols = [col for col in monthly_room_sales.columns if col != '객실']
    monthly_room_sales['합계'] = monthly_room_sales[month_cols].sum(axis=1)
    
    # Add column total (sum for each month across all rooms) - 마지막 행에 합계
    total_row_data = ['합계'] + [monthly_room_sales[col].sum() for col in month_cols] + [monthly_room_sales['합계'].sum()]
    total_row = pd.DataFrame([total_row_data], columns=monthly_room_sales.columns)
    monthly_room_sales = pd.concat([monthly_room_sales, total_row], ignore_index=True)

    # 4. Daily Occupancy & Room Type Stats
    room_mapping = {
        '302': '이코노미더블룸', '305': '이코노미더블룸', '306': '이코노미더블룸',
        '201': '더블룸', '401': '더블룸', '501': '더블룸', '601': '더블룸', '701': '더블룸',
        '205': '싱글룸', '206': '싱글룸', '405': '싱글룸', '406': '싱글룸',
        '505': '싱글룸', '506': '싱글룸', '605': '싱글룸', '606': '싱글룸',
        '705': '싱글룸', '706': '싱글룸',
        '203': '트윈룸', '204': '트윈룸', '303': '트윈룸', '304': '트윈룸',
        '403': '트윈룸', '404': '트윈룸', '503': '트윈룸', '504': '트윈룸',
        '603': '트윈룸', '604': '트윈룸', '703': '트윈룸', '704': '트윈룸',
        '202': '트리플룸', '402': '트리플룸', '502': '트리플룸', 
        '602': '트리플룸', '702': '트리플룸', '803': '트리플룸',
        '801': '패밀리4',
        '802': '패밀리5'
    }
    
    room_capacities = {
        '이코노미더블룸': 3,
        '더블룸': 5,
        '싱글룸': 10,
        '트윈룸': 12,
        '트리플룸': 6,
        '패밀리4': 1,
        '패밀리5': 1
    }

    def get_room_type(room):
        try:
            # Normalize room number: '0202' -> 202, 202 -> 202
            normalized_room = str(int(str(room)))
        except (ValueError, TypeError):
            normalized_room = str(room).strip()
            
        return room_mapping.get(normalized_room, '기타')

    df['객실타입'] = df['객실'].apply(get_room_type)
    
    # 4.1 Daily Stats (Total)
    daily_total = df.groupby('일자').agg(
        총판매객실수=('객실', 'nunique'),
        전체평균가격=('총금액', 'mean')
    ).reset_index()
    daily_total['전체점유율(%)'] = (daily_total['총판매객실수'] / 38) * 100

    # 4.2 Daily Stats by Room Type (Pivot)
    # Pivot for Counts
    daily_counts = df.pivot_table(index='일자', columns='객실타입', values='객실', aggfunc='nunique', fill_value=0)
    # Pivot for Average Prices
    daily_prices = df.pivot_table(index='일자', columns='객실타입', values='총금액', aggfunc='mean', fill_value=0)

    # Merge everything into one DataFrame
    daily_report = daily_total.copy()
    daily_report.set_index('일자', inplace=True)

    # Define order of room types for display
    room_order = ['싱글룸', '이코노미더블룸', '더블룸', '트윈룸', '트리플룸', '패밀리4', '패밀리5']

    for r_type in room_order:
        capacity = room_capacities.get(r_type, 0)
        
        # Count
        if r_type in daily_counts.columns:
            col_count = daily_counts[r_type]
        else:
            col_count = 0
        
        # Price
        if r_type in daily_prices.columns:
            col_price = daily_prices[r_type]
        else:
            col_price = 0
            
        # Calculate %
        if capacity > 0:
            col_pct = (col_count / capacity) * 100
        else:
            col_pct = 0
        
        # Add to report with clear column names
        daily_report[f'{r_type}_판매수'] = col_count
        daily_report[f'{r_type}_점유율(%)'] = col_pct
        daily_report[f'{r_type}_평균가격'] = col_price

    daily_report.reset_index(inplace=True)
    
    # Add Day of Week
    days = ['월', '화', '수', '목', '금', '토', '일']
    daily_report['요일'] = daily_report['일자'].dt.dayofweek.apply(lambda x: days[x])
    
    # Reorder columns to put '요일' next to '일자'
    cols = daily_report.columns.tolist()
    # cols[0] is '일자', insert '요일' at index 1
    if '요일' in cols:
        cols.insert(1, cols.pop(cols.index('요일')))
    daily_report = daily_report[cols]
    
    daily_report['일자'] = daily_report['일자'].dt.strftime('%Y-%m-%d')

    # 4.3 Room Type Stats (Overall Aggregated)
    room_type_stats = df.groupby('객실타입')['총금액'].agg(['mean', 'sum', 'count']).reset_index()
    room_type_stats.columns = ['객실타입', '평균가격', '총매출', '판매건수']
    
    # 4.4 Overall Average Price
    overall_avg_price = df['총금액'].mean()
    overall_stats = pd.DataFrame({'구분': ['전체 평균가격'], '값': [overall_avg_price]})

    # Load existing workbook
    wb = openpyxl.load_workbook(file_path)
    
    # Helper function to add sheet if not exists and write data
    def write_dataframe_to_sheet(workbook, sheet_name, dataframe):
        if sheet_name in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' already exists. Skipping.")
            return workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)
            for r in dataframe_to_rows(dataframe, index=False, header=True):
                ws.append(r)
            print(f"Created sheet '{sheet_name}'.")
            return ws

    # Add sheets
    ws_monthly = write_dataframe_to_sheet(wb, '월별매출', monthly_sales)
    ws_room = write_dataframe_to_sheet(wb, '룸별매출', room_sales)
    ws_monthly_room = write_dataframe_to_sheet(wb, '월별_룸별매출', monthly_room_sales)

    # New Sheet: Occupancy and Average Price
    sheet_name_new = '점유율_및_평균단가'
    
    # Remove existing sheets that start with the target name (including v2, v3, etc.)
    sheets_to_remove = [s for s in wb.sheetnames if s.startswith('점유율_및_평균단가')]
    for s in sheets_to_remove:
        del wb[s]
        print(f"Removed existing sheet '{s}'.")
        
    ws_new = wb.create_sheet(sheet_name_new)
    
    # Write Daily Report (Detailed)
    ws_new.append(['[일별 상세 현황 (판매수, 점유율, 평균가격)]'])
    for r in dataframe_to_rows(daily_report, index=False, header=True):
        ws_new.append(r)
        
    ws_new.append([]) # Empty row
    ws_new.append(['[룸별 전체 통계]'])
    
    # Write Room Type Stats
    for r in dataframe_to_rows(room_type_stats, index=False, header=True):
        ws_new.append(r)

    ws_new.append([]) # Empty row
    ws_new.append(['[전체 통계]'])
    for r in dataframe_to_rows(overall_stats, index=False, header=True):
        ws_new.append(r)
        
    print(f"Created sheet '{sheet_name_new}'.")

    # Formatting function
    def apply_styles(ws, sheet_type='default'):
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        comma_format = '#,##0'
        
        # For simple sales sheets (월별매출, 룸별매출, 월별_룸별매출)
        if sheet_type == 'sales':
            for row in ws.iter_rows(min_row=1):
                for cell in row:
                    cell.border = thin_border
                    # Apply comma format to numeric values
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = comma_format
            return
        
        # Define fills for conditional formatting
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        
        # Define font for holidays/weekends
        red_bold_font = Font(color='FF0000', bold=True)
        
        # Holidays list (2025)
        holidays = [
            '2025-01-01', '2025-01-28', '2025-01-29', '2025-01-30',
            '2025-03-01', '2025-03-03', '2025-05-05', '2025-05-06',
            '2025-06-06', '2025-08-15', '2025-10-03', '2025-10-05',
            '2025-10-06', '2025-10-07', '2025-10-09', '2025-12-25'
        ]
        
        # Identify columns
        occupancy_cols = []
        date_col_idx = None
        day_col_idx = None
        
        header_row = ws[2] 
        for cell in header_row:
            val_str = str(cell.value)
            if cell.value == '일자':
                date_col_idx = cell.column
            elif cell.value == '요일':
                day_col_idx = cell.column
            elif cell.value and '점유율(%)' in val_str:
                occupancy_cols.append(cell.column)

        for row in ws.iter_rows(min_row=3): # Start from data row
            # Check for holiday/weekend
            is_special_day = False
            date_val = None
            day_val = None
            
            if date_col_idx:
                date_val = row[date_col_idx-1].value
            if day_col_idx:
                day_val = row[day_col_idx-1].value
                
            if day_val in ['금', '토']:
                is_special_day = True
            
            # Handle date comparison safely
            if date_val:
                date_str = str(date_val)
                if hasattr(date_val, 'strftime'):
                    date_str = date_val.strftime('%Y-%m-%d')
                elif ' ' in date_str: # Handle string with time '2025-01-01 00:00:00'
                    date_str = date_str.split(' ')[0]
                
                if date_str in holidays:
                    is_special_day = True
                
            for cell in row:
                cell.border = thin_border
                # Apply comma format to numeric values (excluding header)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = comma_format
                    
                    # Conditional formatting for occupancy columns
                    if cell.column in occupancy_cols:
                        if 80 <= cell.value < 100: # 80% or more but less than 100%
                            cell.fill = yellow_fill
                        elif 51 <= cell.value < 80: # 51% to just under 80%
                            cell.fill = orange_fill
                        elif cell.value <= 50: # 50% or less
                            cell.fill = red_fill
                            
                # Apply Red Bold to Date and Day columns
                if is_special_day:
                    if cell.column == date_col_idx or cell.column == day_col_idx:
                        cell.font = red_bold_font

    # Apply styles to new sheets
    if '월별매출' in wb.sheetnames: apply_styles(wb['월별매출'], sheet_type='sales')
    if '룸별매출' in wb.sheetnames: apply_styles(wb['룸별매출'], sheet_type='sales')
    if '월별_룸별매출' in wb.sheetnames: apply_styles(wb['월별_룸별매출'], sheet_type='sales')
    if '점유율_및_평균단가' in wb.sheetnames: apply_styles(wb['점유율_및_평균단가'])

    wb.save(output_file_path)
    print(f"Analysis completed and saved to {output_file_path}")

except Exception as e:
    print(f"An error occurred: {e}")
